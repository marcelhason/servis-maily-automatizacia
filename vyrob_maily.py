"""
Vyrobi .eml subory zo vzoru `VZOR - ...Vzorová Firma s.r.o..eml`
pre vsetky servisy z `unikaty_podla_H.xlsx`, ktorym priradime adekvatnu
prilohu z `servis_subory/`.

Spustenie testu (len 3 servisy):
    python vyrob_maily.py --test

Spustenie pre vsetky servisy s mailom:
    python vyrob_maily.py
"""
import base64
import datetime as dt
import quopri
import random
import re
import sys
from collections import defaultdict
from email.header import Header
from pathlib import Path

import openpyxl

# Osobne udaje do podpisu beru z lokalneho config.py (gitignored); ak chyba,
# fallback na vzorove placeholder hodnoty z config_example.py.
try:
    import config as _cfg
except ImportError:
    import config_example as _cfg

ROOT = Path(__file__).resolve().parent
VZOR = ROOT / "VZOR - Žiadosť o informácie k doriešeniu reklamácií  – Vzorová Firma s.r.o..eml"
UNIKATY = ROOT / "unikaty_podla_H.xlsx"
SERVIS_DIR = ROOT / "servis_subory"
OUT_DIR = ROOT / "maily_servisom"

SUBJECT_PREFIX = "Žiadosť o informácie k doriešeniu reklamácií  – "

INVALID_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|]')


# ---------- pomocne funkcie ----------

def sanitize_filename(name: str) -> str:
    return INVALID_FILENAME_CHARS.sub("_", name).strip()


def extract_service_name(h: str) -> str:
    return h.split(" - ", 1)[1].strip() if " - " in h else h.strip()


def extract_service_code(h: str) -> str:
    return h.split(" - ", 1)[0].strip() if " - " in h else ""


def fmt_date(v) -> str:
    if isinstance(v, (dt.datetime, dt.date)):
        return f"{v.day}.{v.month}.{v.year}"
    return str(v) if v is not None else ""


# ---------- nacitanie unikatov ----------

def load_unikaty():
    """Vrati list dict: {h, email1, email2, nazov}. Iba zaznamy s email1."""
    wb = openpyxl.load_workbook(UNIKATY, data_only=True)
    ws = wb.active
    seen = set()
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        h, email1, email2 = row[3], row[1], row[2]
        if not h or not email1:
            continue
        if h in seen:
            continue
        seen.add(h)
        out.append({
            "h": h,
            "email1": str(email1).strip(),
            "email2": str(email2).strip() if email2 else None,
            "nazov": extract_service_name(h),
        })
    return out


def build_filename_map(records):
    """h -> nazov suboru bez pripony (s priponou (kod) pri kolizii nazvov)."""
    by_name = defaultdict(list)
    for r in records:
        by_name[r["nazov"]].append(r)
    out = {}
    for name, recs in by_name.items():
        safe = sanitize_filename(name)
        if len(recs) == 1:
            out[recs[0]["h"]] = safe
        else:
            for r in recs:
                out[r["h"]] = f"{safe} ({extract_service_code(r['h'])})"
    return out


def read_xlsx_rows(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        rows.append(row)
    return rows


# ---------- generovanie tela ----------

PLAIN_INTRO = (
    "Dobry den,\r\n\r\n\r\n\r\n"
    "Obraciam sa na Vas s prosbou o preverenie a zaslanie informacii o stave doriesenia nasich reklamacii.\r\n\r\n"
    "Vsetky potrebne identifikacne udaje k jednotlivym pripadom najdete v prilozenej Excel tabulke.\r\n\r\n\r\n\r\n"
    "Pri kazdej RMA by sme potrebovali vediet, ako bola definitivne uzavreta. Prosim Vas o poskytnutie nasledovnych informacii:\r\n\r\n\r\n\r\n"
    "1.      Ak bola reklamacia doriesena dobropisom:\r\n\r\n"
    "   *    Prosime o zaslanie potvrdenia o dobropise alebo o uvedenie cisla dobropisu.\r\n\r\n\r\n\r\n"
    "2.      Ak reklamacia presla opravou:\r\n\r\n"
    "   *    Potrebujeme informaciu, kedy a na aku adresu bola opravena RMA vratena, idealne aj s potvrdenim o doruceni/odoslani.\r\n\r\n\r\n\r\n"
    "   Ak je to pre Vas jednoduchsie, informacie mozete dopisat priamo do prilozenej tabulky a poslat nam ju spat spolu s potrebnymi dokumentmi v prilohe.\r\n\r\n\r\n\r\n"
    "   Vopred Vam velmi pekne dakujem za spolupracu a rychlu odpoved.\r\n\r\n\r\n\r\n\r\n\r\n"
)
PLAIN_OUTRO_HEADERS = ["Hlášení", "Krát.text materiálu", "Název značky",
                       "Servis", "Datum hlášení", "Poznámka servis"]
PLAIN_FOOTER = (
    "\r\n\r\n\r\n\r\n\r\n\r\n"
    "S pozdravom\r\n\r\n\r\n\r\n"
    f"{_cfg.SIGNATURE_NAME}\r\n\r\n"
    f"{_cfg.SIGNATURE_ROLE}\r\n\r\n\r\n\r\n"
    + "\r\n\r\n".join(_cfg.SIGNATURE_COMPANY_LINES)
    + "\r\n\r\n\r\n\r\n\r\n\r\n"
)


def build_plain_body(rows) -> str:
    """Plain text telo (este nezakodovane do quoted-printable)."""
    parts = [PLAIN_INTRO]
    # hlavicka tabulky - kazde pole na vlastnom riadku, prazdny riadok medzi
    for h in PLAIN_OUTRO_HEADERS:
        parts.append(h + "\r\n\r\n")
    parts.append("\r\n")  # extra prazdny riadok pred datami (vo vzore je extra blank)
    for row in rows:
        hlaseni, ktext, znacka, servis, datum, poznamka = row
        cells = [
            str(hlaseni) if hlaseni is not None else "",
            str(ktext) if ktext else "",
            str(znacka) if znacka else "",
            str(servis) if servis else "",
            fmt_date(datum),
            str(poznamka) if poznamka else "",
        ]
        for c in cells:
            parts.append(c + "\r\n\r\n")
        parts.append("\r\n")  # extra blank po riadku
    parts.append(PLAIN_FOOTER)
    return "".join(parts)


# ---------- HTML tabulka ----------

HTML_TR_HEADER_TEMPLATE = """<tr style="height:15.0pt">
<td width="84" nowrap valign="top" style="width:63.0pt;border:solid windowtext 1.0pt;background:#BFBFBF;padding:0cm 3.5pt 0cm 3.5pt;height:15.0pt"><p class="MsoNormal"><span style="font-size:11.0pt;font-family:&quot;Aptos Narrow&quot;,sans-serif;color:black;mso-ligatures:none;mso-fareast-language:SK">Hlášení<o:p></o:p></span></p></td>
<td width="295" nowrap valign="top" style="width:221.0pt;border:solid windowtext 1.0pt;border-left:none;background:#BFBFBF;padding:0cm 3.5pt 0cm 3.5pt;height:15.0pt"><p class="MsoNormal"><span style="font-size:11.0pt;font-family:&quot;Aptos Narrow&quot;,sans-serif;color:black;mso-ligatures:none;mso-fareast-language:SK">Krát.text materiálu<o:p></o:p></span></p></td>
<td width="125" nowrap valign="top" style="width:94.0pt;border:solid windowtext 1.0pt;border-left:none;background:#BFBFBF;padding:0cm 3.5pt 0cm 3.5pt;height:15.0pt"><p class="MsoNormal"><span style="font-size:11.0pt;font-family:&quot;Aptos Narrow&quot;,sans-serif;color:black;mso-ligatures:none;mso-fareast-language:SK">Název značky<o:p></o:p></span></p></td>
<td width="357" nowrap valign="top" style="width:268.0pt;border:solid windowtext 1.0pt;border-left:none;background:#BFBFBF;padding:0cm 3.5pt 0cm 3.5pt;height:15.0pt"><p class="MsoNormal"><span style="font-size:11.0pt;font-family:&quot;Aptos Narrow&quot;,sans-serif;color:black;mso-ligatures:none;mso-fareast-language:SK">Servis<o:p></o:p></span></p></td>
<td width="105" nowrap valign="top" style="width:79.0pt;border:solid windowtext 1.0pt;border-left:none;background:#BFBFBF;padding:0cm 3.5pt 0cm 3.5pt;height:15.0pt"><p class="MsoNormal"><span style="font-size:11.0pt;font-family:&quot;Aptos Narrow&quot;,sans-serif;color:black;mso-ligatures:none;mso-fareast-language:SK">Datum hlášení<o:p></o:p></span></p></td>
<td width="353" nowrap valign="bottom" style="width:265.0pt;border:solid windowtext 1.0pt;border-left:none;background:#BFBFBF;padding:0cm 3.5pt 0cm 3.5pt;height:15.0pt"><p class="MsoNormal"><span style="font-size:11.0pt;font-family:&quot;Aptos Narrow&quot;,sans-serif;color:black;mso-ligatures:none;mso-fareast-language:SK">Poznámka servis<o:p></o:p></span></p></td>
</tr>"""


def _cell(width_attr, width_style, content, is_last_row, align_right=False, is_last_col=False):
    """Postavi <td> bunku. is_last_row pridava border-bottom."""
    border = ""
    if is_last_row:
        if not align_right and width_attr == "84":  # prvy stlpec
            border = "border-top:none;border-left:solid windowtext 1.0pt;border-bottom:solid windowtext 1.0pt;border-right:none;"
        elif is_last_col:
            border = "border:none;border-bottom:solid windowtext 1.0pt;"
        else:
            border = "border:none;border-bottom:solid windowtext 1.0pt;"
    height = "15.75pt" if is_last_row else "15.0pt"
    style = f'width:{width_style};{border}padding:0cm 3.5pt 0cm 3.5pt;height:{height}'
    # prvy stlpec ma vzdy lavu hranicu (aj v normalnych riadkoch)
    if not is_last_row and width_attr == "84":
        style = f'width:{width_style};border:none;border-left:solid windowtext 1.0pt;padding:0cm 3.5pt 0cm 3.5pt;height:15.0pt'
    p_attrs = ' align="right" style="text-align:right"' if align_right else ""
    content_html = content if content else "&nbsp;"
    return (f'<td width="{width_attr}" nowrap valign="top" style="{style}">'
            f'<p class="MsoNormal"{p_attrs}>'
            f'<span style="font-size:11.0pt;font-family:&quot;Aptos Narrow&quot;,sans-serif;'
            f'color:black;mso-ligatures:none;mso-fareast-language:SK">'
            f'{content_html}<o:p></o:p></span></p></td>')


def build_html_data_rows(rows):
    """Vrati string s <tr>...</tr> riadkami pre data (bez hlavickoveho riadku)."""
    out = []
    last_idx = len(rows) - 1
    cols = [("84", "63.0pt"), ("295", "221.0pt"), ("125", "94.0pt"),
            ("357", "268.0pt"), ("105", "79.0pt"), ("353", "265.0pt")]
    for i, row in enumerate(rows):
        is_last = (i == last_idx)
        height = "15.75pt" if is_last else "15.0pt"
        hlaseni, ktext, znacka, servis, datum, poznamka = row
        values = [
            str(hlaseni) if hlaseni is not None else "",
            (str(ktext) if ktext else "").replace("&", "&amp;"),
            (str(znacka) if znacka else "").replace("&", "&amp;"),
            (str(servis) if servis else "").replace("&", "&amp;"),
            fmt_date(datum),
            (str(poznamka) if poznamka else ""),
        ]
        cells_html = []
        for j, ((wa, ws), val) in enumerate(zip(cols, values)):
            align_right = (j == 4)  # datum
            is_last_col = (j == 5)
            cells_html.append(_cell(wa, ws, val, is_last, align_right=align_right, is_last_col=is_last_col))
        out.append(f'<tr style="height:{height}">\n' + "\n".join(cells_html) + "\n</tr>")
    return "\n".join(out)


# ---------- kodovanie ----------

def to_quoted_printable_cp1250(text: str) -> bytes:
    """Zakoduje string do cp1250 + quoted-printable s CRLF line breakmi."""
    raw = text.encode("cp1250")
    # quopri encodestring rozdeli na linky a pripoji "=" pre soft breaks
    qp = quopri.encodestring(raw, quotetabs=False)
    # quopri pouziva \n pre line endings - prerobime na \r\n
    qp = qp.replace(b"\r\n", b"\n").replace(b"\n", b"\r\n")
    return qp


def encode_header_q(text: str) -> bytes:
    """Q-encoded hlavicka v windows-1250 (multiline s ' ' continuation)."""
    h = Header(text, "windows-1250")
    enc = h.encode(maxlinelen=72)
    # email.header pouziva \n - prerobime na \r\n a continuation ('\n ' -> '\r\n ')
    lines = enc.split("\n")
    return ("\r\n ".join(l.lstrip() for l in lines)).encode("ascii")


def q_encode_filename(name: str) -> str:
    """Pre name v Content-Type / filename hlavickach. Bez line wrappingu."""
    raw = name.encode("cp1250")
    out = []
    for b in raw:
        if b == 0x20:
            out.append("_")
        elif 0x30 <= b <= 0x39 or 0x41 <= b <= 0x5A or 0x61 <= b <= 0x7A or b in (0x2D, 0x2E, 0x2F):
            out.append(chr(b))
        else:
            out.append(f"={b:02X}")
    return "=?windows-1250?Q?" + "".join(out) + "?="


# ---------- konstrukcia .eml ----------

BOUNDARY_006 = "_006_GVUPR03MB117227ECD1EE70B5F8DE9D33EC11D2GVUPR03MB11722eu_"
BOUNDARY_005 = "_005_GVUPR03MB117227ECD1EE70B5F8DE9D33EC11D2GVUPR03MB11722eu_"
BOUNDARY_000 = "_000_GVUPR03MB117227ECD1EE70B5F8DE9D33EC11D2GVUPR03MB11722eu_"


def extract_image_block(vzor_bytes: bytes) -> bytes:
    """Vytiahne kompletny image001.png blok zo vzoru (od --_005_ s image az do nasledujuceho boundary)."""
    start = vzor_bytes.index(b"--" + BOUNDARY_005.encode() + b"\r\nContent-Type: image/png")
    # koniec = pozicia naslednehu '--BOUNDARY_005--'
    end = vzor_bytes.index(b"--" + BOUNDARY_005.encode() + b"--", start)
    return vzor_bytes[start:end]  # vratane uvodneho boundary, bez zaverecneho


def extract_html_template(vzor_bytes: bytes):
    """Vrati (html_pred_tabulkou, html_po_tabulke) - vsetko v cp1250+qp.
    Tabulkova hlavicka (header row) je zachovana ako sucast 'pred'."""
    # najdi text/html sekciu
    html_start_marker = b"--" + BOUNDARY_000.encode() + b"\r\nContent-Type: text/html"
    html_start = vzor_bytes.index(html_start_marker)
    # konkretne telo zacina po prazdnom riadku po headeroch tohto blocku
    body_start = vzor_bytes.index(b"\r\n\r\n", html_start) + 4
    body_end = vzor_bytes.index(b"\r\n--" + BOUNDARY_000.encode(), body_start)
    return vzor_bytes[html_start:body_start], vzor_bytes[body_start:body_end]


def extract_plain_template(vzor_bytes: bytes):
    """Vrati (plain_blok_header_a_telo) - cely text/plain blok aj s boundary za nim NEbude."""
    plain_start_marker = b"--" + BOUNDARY_000.encode() + b"\r\nContent-Type: text/plain"
    plain_start = vzor_bytes.index(plain_start_marker)
    body_start = vzor_bytes.index(b"\r\n\r\n", plain_start) + 4
    body_end = vzor_bytes.index(b"\r\n--" + BOUNDARY_000.encode(), body_start)
    headers = vzor_bytes[plain_start:body_start]
    return headers, vzor_bytes[body_start:body_end]


def build_new_html_body(vzor_bytes: bytes, rows) -> bytes:
    """Postavi text/html telo s novymi datovymi riadkami tabulky."""
    _, html_body_orig = extract_html_template(vzor_bytes)
    # dekoduj cp1250 qp -> str (aby sme mohli textovo nahradit)
    html_text = quopri.decodestring(html_body_orig).decode("cp1250")
    # najdi tabulku (zacina '<table class="MsoNormalTable"') a obsah medzi prvym </tr> a </tbody>
    m = re.search(r'(<table class="MsoNormalTable"[\s\S]*?</tr>)([\s\S]*?)(</tbody>)', html_text)
    if not m:
        raise RuntimeError("Tabulku v HTML sa nepodarilo najst.")
    new_rows_html = build_html_data_rows(rows)
    html_text = html_text[:m.end(1)] + "\n" + new_rows_html + "\n" + html_text[m.start(3):]
    return to_quoted_printable_cp1250(html_text)


def build_new_plain_body(rows) -> bytes:
    plain_text = build_plain_body(rows)
    return to_quoted_printable_cp1250(plain_text)


def build_attachment_block(filename: str, file_bytes: bytes, now: dt.datetime) -> bytes:
    """Postavi attachment cast vratane uvodneho --_006_ a Headerov + base64 obsahu.
    Nezahrnut konecne --_006_--."""
    qname = q_encode_filename(filename)
    b64 = base64.encodebytes(file_bytes)  # vlozi \n kazdych 76 znakov
    b64 = b64.replace(b"\r\n", b"\n").replace(b"\n", b"\r\n")
    date_str = now.strftime("%a, %d %b %Y %H:%M:%S GMT")
    headers = (
        f"--{BOUNDARY_006}\r\n"
        f"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;\r\n"
        f"\tname=\"{qname}\"\r\n"
        f"Content-Description: {qname}\r\n"
        f"Content-Disposition: attachment;\r\n"
        f"\tfilename=\"{qname}\"; size={len(file_bytes)};\r\n"
        f"\tcreation-date=\"{date_str}\";\r\n"
        f"\tmodification-date=\"{date_str}\"\r\n"
        f"Content-Transfer-Encoding: base64\r\n\r\n"
    ).encode("ascii")
    return headers + b64


def build_eml(vzor_bytes: bytes, *, subject: str, to_addr: str, rows,
              attachment_name: str, attachment_bytes: bytes) -> bytes:
    """Postavi cely novy .eml ako bytes."""
    # 1) MINIMALNE hlavicky pre Outlook draft - X-Unsent: 1 pred Content-Type,
    # bez From/Date/Message-ID/X-Mailer (Outlook si doplni From z aktivneho uctu)
    subject_enc = encode_header_q(subject)
    headers = (
        b"X-Unsent: 1\r\n"
        b"To: " + to_addr.encode("ascii") + b"\r\n"
        b"Subject: " + subject_enc + b"\r\n"
        b"MIME-Version: 1.0\r\n"
        b"Content-Type: multipart/mixed;\r\n"
        b"\tboundary=\"" + BOUNDARY_006.encode() + b"\"\r\n"
    )

    # 2) preamble
    preamble = b"\r\n"

    # 3) prva cast multipart/mixed = multipart/related
    related_open = (
        f"--{BOUNDARY_006}\r\n"
        f"Content-Type: multipart/related;\r\n"
        f"\tboundary=\"{BOUNDARY_005}\";\r\n"
        f"\ttype=\"multipart/alternative\"\r\n\r\n"
    ).encode("ascii")

    # 3a) vnutri related: multipart/alternative
    alt_open = (
        f"--{BOUNDARY_005}\r\n"
        f"Content-Type: multipart/alternative;\r\n"
        f"\tboundary=\"{BOUNDARY_000}\"\r\n\r\n"
    ).encode("ascii")

    # 3a-i) text/plain
    plain_headers = (
        f"--{BOUNDARY_000}\r\n"
        f"Content-Type: text/plain; charset=\"windows-1250\"\r\n"
        f"Content-Transfer-Encoding: quoted-printable\r\n\r\n"
    ).encode("ascii")
    plain_body = build_new_plain_body(rows)

    # 3a-ii) text/html
    html_headers = (
        f"--{BOUNDARY_000}\r\n"
        f"Content-Type: text/html; charset=\"windows-1250\"\r\n"
        f"Content-Transfer-Encoding: quoted-printable\r\n\r\n"
    ).encode("ascii")
    html_body = build_new_html_body(vzor_bytes, rows)

    alt_close = f"--{BOUNDARY_000}--\r\n".encode("ascii")

    # 3b) v related: obrazok (kopia zo vzoru)
    image_block = extract_image_block(vzor_bytes)
    related_close = f"--{BOUNDARY_005}--\r\n".encode("ascii")

    # 4) druha cast multipart/mixed = attachment
    attach_block = build_attachment_block(attachment_name, attachment_bytes, dt.datetime.now())

    mixed_close = f"--{BOUNDARY_006}--\r\n".encode("ascii")

    return (
        headers + preamble +
        related_open +
        alt_open +
        plain_headers + plain_body + b"\r\n" +
        html_headers + html_body + b"\r\n" +
        alt_close + b"\r\n" +
        image_block +
        related_close + b"\r\n" +
        attach_block + b"\r\n" +
        mixed_close
    )


# ---------- main ----------

def main():
    test_mode = "--test" in sys.argv
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    records = load_unikaty()
    fname_map = build_filename_map(records)

    if test_mode:
        # 1 nahodny zaznam s emailom
        targets = random.sample(records, 1)
    else:
        targets = records

    vzor_bytes = VZOR.read_bytes()

    print(f"Spracovavam {len(targets)} servisov (test_mode={test_mode}, mam emailov={len(records)})")
    skipped = []
    ok = 0
    for rec in targets:
        h = rec["h"]
        nazov = rec["nazov"]
        base_fname = fname_map[h]
        xlsx_path = SERVIS_DIR / f"{base_fname}.xlsx"
        if not xlsx_path.exists():
            print(f"  SKIP (chyba xlsx): {xlsx_path.name}")
            skipped.append(base_fname)
            continue

        rows = read_xlsx_rows(xlsx_path)
        attachment_bytes = xlsx_path.read_bytes()
        attachment_name = f"{base_fname}.xlsx"

        subject = SUBJECT_PREFIX + nazov
        to_parts = [rec["email1"]]
        if rec["email2"]:
            to_parts.append(rec["email2"])
        to_addr = ", ".join(to_parts)

        new_eml = build_eml(
            vzor_bytes,
            subject=subject,
            to_addr=to_addr,
            rows=rows,
            attachment_name=attachment_name,
            attachment_bytes=attachment_bytes,
        )

        out_name = sanitize_filename(f"{SUBJECT_PREFIX}{nazov}") + ".eml"
        out_path = OUT_DIR / out_name
        out_path.write_bytes(new_eml)
        ok += 1
        print(f"  OK {out_name}  ({len(rows)} riadkov, {len(attachment_bytes)} B priloha, to={to_addr})")

    print(f"\nHotovo: {ok} OK, {len(skipped)} preskocenych")
    print(f"Vystup: {OUT_DIR}")


if __name__ == "__main__":
    main()
