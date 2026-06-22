"""
Spracuje stiahnuté odpovede (odpovede/<mail_id>/) a vygeneruje výstupy
pre človeka: strom priečinkov podľa servisov + excel s plnou stopou.

Postup per mail:
  1. spáruje mail so servisom (názov v predmete; fallback známy email),
  2. prečíta ho DVOMA vrstvami: heuristika (parsers.parse_reply +
     parse_xlsx_reply) a AI (parsers.parse_reply_ai, ak je v config.py
     ANTHROPIC_API_KEY); výsledky zlúči parsers.zluc_ai_heur —
     len-AI nález pod prahom AI_MIN_CONFIDENCE sa NEzapíše a mail ide
     na ručnú kontrolu aj s AI návrhmi,
  3. zapíše append-only do stav_reklamacii.json (dedup podľa mail_id+zdroj),
  4. AI odpovede sa cachujú v odpovede/<mail_id>/ai_vysledok.json —
     opakovaný beh nevolá API (vynúti --ai-znova).

Výstupy (vždy sa pregenerujú nanovo, sú to len pohľady na stav + cache):
  odpovede_podla_servisov/<SERVIS>/   čitateľné maily (.txt s hlavičkou
                                      Od/Prijaté/Predmet) a prílohy
                                      s dátumom v názve — na SAP zálohu
                                      stačí skopírovať priečinok servisu
  dpb servis 180-360 - stav.xlsx      stĺpce Odpoveď servisu / Od koho /
                                      Predmet mailu / odkazy na priečinok
                                      + hárok „Prehľad servisov"

Excel + strom priečinkov patria k sebe (relatívne odkazy) — Eve sa
kopírujú SPOLU do jedného priečinka.

Tretia vrstva (klasifikácia): Claude Code agent klasifikator-odpovedi
(spúšťa sa cez /klasifikuj-odpovede) zapisuje odpovede/<id>/klasifikacia.json
— kategória reklamácie, akcia a hotový „pokec" pre SAP. Tento skript ich
merguje do stavu (zdroj "agent") a kreslí stĺpce Kategória / Číslo
dobropisu / Akcia / Pokec pre SAP + hárok „Triáž" (pracovný zoznam
zoskupený podľa akcií).

Použitie:
    python spracuj_odpovede.py                  # spracuj + strom + excel
    python spracuj_odpovede.py --len-excel      # bez spracovania (bez AI)
    python spracuj_odpovede.py --len-priecinky  # iba strom priečinkov
    python spracuj_odpovede.py --ai-znova       # ignoruj AI cache
    python spracuj_odpovede.py --len-klasifikacie  # merge klasifikácií
                                                   # (bez AI vrstvy) + výstupy
"""
import argparse
import json
import re
import shutil
import sys
import urllib.parse
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).parent))
from vyrob_maily import ROOT  # noqa: E402
from rozdel_servisy import SRC  # noqa: E402
from parsers import (  # noqa: E402
    ai_dostupne, eml_na_text, parse_reply, parse_reply_ai, parse_xlsx_reply,
    xlsx_na_text, zluc_ai_heur, ParsedItem, _norm,
    CONFIDENCE_PORADIE, HLASENIE_RE,
)
import stav as stavmod  # noqa: E402

ODPOVEDE_DIR = ROOT / "odpovede"
STROM_DIR = ROOT / "odpovede_podla_servisov"
PRILOHY_DIR = ROOT / "prilohy"
MAILY_DIR = ROOT / "maily"
VYSTUP_XLSX = ROOT / "dpb servis 180-360 - stav.xlsx"

# Vzor deep linku na konverzáciu v OWA: po otvorení mailu má URL tvar
# .../mail/<id priečinka>/id/<convid url-encoded>. Scraper ukladá hotovú
# owa_url do meta.json; pre maily bez nej slúži vzor z config.py
# (OWA_LINK_VZOR, {id} = url-encoded convid). POZOR: otvorenie deep linku
# beží mimo filtra Neprečítané, takže mail označí za prečítaný — pri
# vybavovaní Evou je to žiaduce.
try:
    from config import OWA_LINK_VZOR  # noqa: E402
except ImportError:
    OWA_LINK_VZOR = ""

NOVE_HLAVICKY = ["Stav", "Na ťahu", "Odpoveď servisu", "Dátum odpovede",
                 "Od koho", "Predmet mailu", "Mail v Outlooku", "Priečinok",
                 "História", "Ručná kontrola",
                 "Kategória", "Číslo dobropisu", "Akcia", "Pokec pre SAP",
                 "Prílohy"]

# zdroje poznámok, ktoré sú odpoveďou servisu (nie ziadost/manual)
ODPOVED_ZDROJE = ("telo", "priloha", "ai", "ai+heur", "agent")
ZDROJ_POPIS = {"telo": "telo mailu", "priloha": "príloha", "ai": "AI",
               "ai+heur": "AI+heuristika", "agent": "klasifikátor",
               "manual": "ručne"}

# enumy klasifikácie — musia sedieť so skillom
# .claude/skills/reklamacie-biznis/SKILL.md; čokoľvek mimo nich sa zahodí
# a mail ide na ručnú kontrolu (žiadne tiché hádanie)
KATEGORIE = ("uzavrete-dobropisovane", "mrtve", "oprava",
             "nerentabilna-oprava", "nedostupnost-nd", "zamietnutie",
             "ping-pong")
MAIL_TYPY = ("vecna-odpoved", "auto-ack", "ndr", "bez-vecnej-informacie")
AKCIE_PORADIE = ("ziadost-o-dobropis-v-sap", "zapisat-do-sap",
                 "pozriet-rucne", "cakat-na-servis")
AKCIA_POPIS = {"ziadost-o-dobropis-v-sap": "Žiadosť o dobropis v SAP",
               "zapisat-do-sap": "Zapísať do SAP",
               "pozriet-rucne": "Pozrieť ručne",
               "cakat-na-servis": "Čakať na servis"}
KATEGORIA_POPIS = {"uzavrete-dobropisovane": "Dobropisované",
                   "mrtve": "Mŕtve (servis nevie)",
                   "oprava": "Opravené",
                   "nerentabilna-oprava": "Nerentabilná oprava",
                   "nedostupnost-nd": "Nedostupný náhradný diel",
                   "zamietnutie": "Zamietnuté",
                   "ping-pong": "Ping-pong"}

HYPERLINK_FONT = Font(color="0563C1", underline="single")
NOVA_FARBA = PatternFill(start_color="FFFF99", end_color="FFFF99",
                         fill_type="solid")


# ---------- párovanie mail -> servis ----------

def najdi_servis(data: dict, meta: dict, telo: str) -> str | None:
    """Vráti servis_h. Primárne podľa názvu servisu v predmete (žiadosti mali
    predmet SUBJECT_PREFIX + nazov a odpovede ho citujú ako RE: ...);
    pri viacerých zhodách vyhrá najdlhší názov (ELEKTROSERVIS vs
    ELEKTROSERVIS PLUS). Fallbacky: email odosielateľa z meta,
    potom známy email servisu v tele mailu."""
    predmet = _norm(meta.get("predmet") or "") + " " \
        + _norm(meta.get("aria_label") or "")
    kandidati = [
        (len(srv["nazov"]), h)
        for h, srv in data["servisy"].items()
        if _norm(srv["nazov"]) in predmet
    ]
    if kandidati:
        return max(kandidati)[1]

    od = (meta.get("odosielatel") or "").lower()
    for h, srv in data["servisy"].items():
        for email in (srv["email1"], srv["email2"]):
            if email and email.lower() in od:
                return h

    telo_n = telo.lower()
    for h, srv in data["servisy"].items():
        for email in (srv["email1"], srv["email2"]):
            if email and email.lower() in telo_n:
                return h
    return None


def je_ndr(meta: dict) -> bool:
    """Nedoručiteľná správa (bounce) od Exchange — nie je to odpoveď servisu,
    znamená mŕtvu adresu."""
    od = meta.get("odosielatel") or ""
    predmet = (meta.get("predmet") or "") + " " + (meta.get("aria_label") or "")
    return ("Microsoft Outlook" in od
            or re.search(r"nedoručiteľn|nedoručiteln|undeliverable",
                         predmet, re.IGNORECASE) is not None)


def odosielatel_casti(meta: dict) -> tuple[str, str]:
    """(meno, email) — z nových polí meta, fallback parse 'Meno<email>'."""
    meno = meta.get("odosielatel_meno") or ""
    email = meta.get("odosielatel_email") or ""
    if not email:
        m = re.match(r"(.*?)<([^<>@\s]+@[^<>\s]+)>", meta.get("odosielatel") or "")
        if m:
            meno, email = m.group(1).strip(), m.group(2).strip()
    return meno, email


def owa_odkaz(meta: dict) -> str | None:
    # Hľadanie podľa odosielateľa + dátum — funguje aj keď Eva mail presunula
    sender = meta.get("odosielatel_email", "")
    datum = meta.get("datum_prijatia", "")   # "2026-06-11"
    if sender and datum:
        q = urllib.parse.quote(f"from:{sender} received:{datum}", safe="")
        return f"https://outlook.cloud.microsoft/mail/search?q={q}"
    return None


def generuj_eml(mail_id: str, meta: dict) -> None:
    from email.mime.text import MIMEText
    from email.header import Header
    from datetime import datetime as _dt

    telo_file = ODPOVEDE_DIR / mail_id / "telo.txt"
    telo = telo_file.read_text(encoding="utf-8", errors="replace") if telo_file.exists() else ""

    msg = MIMEText(telo, "plain", "utf-8")
    msg["From"] = meta.get("odosielatel", "")
    msg["Subject"] = Header(meta.get("predmet", ""), "utf-8")
    datum_str = meta.get("datum_prijatia", "")
    try:
        d = _dt.strptime(datum_str, "%Y-%m-%d")
        msg["Date"] = d.strftime("%a, %d %b %Y 12:00:00 +0000")
    except (ValueError, TypeError):
        pass
    (MAILY_DIR / f"{mail_id}.eml").write_bytes(msg.as_bytes())


def priprav_maily(mapa: dict) -> None:
    shutil.rmtree(MAILY_DIR, ignore_errors=True)
    MAILY_DIR.mkdir(exist_ok=True)
    for mail_id, z in mapa.items():
        generuj_eml(mail_id, z["meta"])
    print(f"Maily (eml): {len(mapa)} → {MAILY_DIR.name}/")


# ---------- mapa stiahnutých mailov (zdieľaná všetkými krokmi) ----------

def zostav_mapu_mailov(data: dict) -> dict[str, dict]:
    """mail_id -> {dir, meta, telo, datum, servis_h, ndr}. Páruje so servisom
    jedenkrát — výsledok zdieľa zápis do stavu, strom priečinkov aj excel."""
    mapa = {}
    if not ODPOVEDE_DIR.exists():
        return mapa
    for mdir in sorted(d for d in ODPOVEDE_DIR.iterdir()
                       if d.is_dir() and (d / "meta.json").exists()):
        meta = json.loads((mdir / "meta.json").read_text(encoding="utf-8"))
        telo = ""
        if (mdir / "telo.txt").exists():
            telo = (mdir / "telo.txt").read_text(encoding="utf-8")
        klasifikacia = None
        if (mdir / "klasifikacia.json").exists():
            try:
                klasifikacia = json.loads(
                    (mdir / "klasifikacia.json").read_text(encoding="utf-8"))
            except json.JSONDecodeError as e:
                print(f"  !! [{mdir.name}] klasifikacia.json sa nedá "
                      f"prečítať: {e}")
        mapa[meta["mail_id"]] = {
            "dir": mdir,
            "meta": meta,
            "telo": telo,
            "datum": meta.get("datum_prijatia") or meta.get("stiahnute") or "",
            "servis_h": najdi_servis(data, meta, telo),
            "ndr": je_ndr(meta),
            "klasifikacia": klasifikacia,
        }
    return mapa


# ---------- AI čítanie s cache ----------

def ai_polozky_mailu(z: dict, zname: set[str],
                     znova: bool = False) -> list[ParsedItem] | None:
    """AI prečíta mail (telo + prepisy xlsx/eml príloh). Výsledok cachuje
    v odpovede/<id>/ai_vysledok.json. None = AI nedostupná/zlyhala —
    caller sa správa, akoby AI vrstva nebola (len heuristika, bez flagov)."""
    cache = z["dir"] / "ai_vysledok.json"
    if cache.exists() and not znova:
        ulozene = json.loads(cache.read_text(encoding="utf-8"))
        return [ParsedItem(**p) for p in ulozene["polozky"]]

    if not ai_dostupne():
        return None

    prepisy = []
    for nazov in z["meta"].get("prilohy", []):
        cesta = z["dir"] / nazov
        if not cesta.exists():
            continue
        try:
            if cesta.suffix.lower() == ".xlsx":
                prepisy.append(f"=== príloha {nazov} ===\n"
                               + xlsx_na_text(cesta))
            elif cesta.suffix.lower() == ".eml":
                prepisy.append(f"=== preposlaný mail {nazov} ===\n"
                               + eml_na_text(cesta))
        except Exception as e:
            print(f"    !! prepis prílohy {nazov}: {type(e).__name__}: {e}")

    if not z["telo"].strip() and not prepisy:
        return None

    try:
        polozky = parse_reply_ai(z["telo"], zname, "\n\n".join(prepisy))
    except Exception as e:
        print(f"    !! AI čítanie zlyhalo: {type(e).__name__}: {e}")
        return None

    import config as _c
    cache.write_text(json.dumps({
        "model": getattr(_c, "AI_MODEL", ""),
        "kedy": date.today().isoformat(),
        "polozky": [vars(p) for p in polozky],
    }, ensure_ascii=False, indent=1), encoding="utf-8")
    return polozky


# ---------- spracovanie do stavového súboru ----------

def spracuj_maily(data: dict, mapa: dict, ai_znova: bool = False) -> dict:
    if not mapa:
        print(f"Priečinok {ODPOVEDE_DIR.name}/ je prázdny — najprv spusti "
              "stiahni_odpovede_owa.py")
        return data
    if not ai_dostupne():
        print("POZN.: ANTHROPIC_API_KEY nie je v config.py — beží len "
              "heuristika, bez AI vrstvy.")

    hlasenia_servisu = {}
    for hid, rec in data["hlasenia"].items():
        hlasenia_servisu.setdefault(rec["servis_h"], set()).add(hid)

    print(f"Spracovávam {len(mapa)} stiahnutých mailov...")

    for mid, z in sorted(mapa.items(), key=lambda kv: (kv[1]["datum"], kv[0])):
        meta, telo, datum = z["meta"], z["telo"], z["datum"]
        servis_h = z["servis_h"]
        if servis_h is None:
            nez = data.setdefault("nezaradene", [])
            if not any(x["mail_id"] == mid for x in nez):
                nez.append({"datum": datum, "mail_id": mid,
                            "predmet": meta.get("predmet") or
                            meta.get("aria_label", "")[:120]})
            print(f"  ?? [{mid}] nespárovaný so servisom")
            continue

        nazov = data["servisy"][servis_h]["nazov"]
        if z["ndr"]:
            ndr = data["servisy"][servis_h].setdefault("nedorucene", [])
            if not any(x["mail_id"] == mid for x in ndr):
                ndr.append({"datum": datum, "mail_id": mid,
                            "predmet": (meta.get("predmet") or "")[:120]})
            print(f"  XX [{mid}] {nazov}: NEDORUČENÉ (mŕtva adresa) — "
                  "neráta sa ako odpoveď")
            continue

        zname = hlasenia_servisu.get(servis_h, set())

        # vrstva 1: heuristika
        heur = []
        for nazov_prilohy in meta.get("prilohy", []):
            cesta = z["dir"] / nazov_prilohy
            if not cesta.exists() or cesta.suffix.lower() != ".xlsx":
                continue
            try:
                heur += [(it, "priloha")
                         for it in parse_xlsx_reply(cesta, zname)]
            except Exception as e:
                print(f"  !! [{mid}] príloha {nazov_prilohy}: "
                      f"{type(e).__name__}: {e}")
        if telo.strip():
            heur += [(it, "telo") for it in parse_reply(telo, zname)]

        # vrstva 2: AI + zlúčenie
        ai_items = ai_polozky_mailu(z, zname, znova=ai_znova)
        if ai_items is None:
            prijate, neiste_ai, len_heur = heur, [], []
        else:
            v = zluc_ai_heur(ai_items, heur)
            prijate, neiste_ai, len_heur = v.prijate, v.neiste_ai, v.len_heur

        zaznamenane = stavmod.zaznamenaj_odpoved_servisu(
            data, servis_h, datum=datum, mail_id=mid)

        pridane, cudzie = 0, 0
        for it, zdroj in prijate:
            if it.id_hlasenia not in data["hlasenia"]:
                cudzie += 1
                continue
            if stavmod.pridaj_odpoved(
                    data, it.id_hlasenia, text=it.status,
                    confidence=it.confidence, datum=datum,
                    zdroj=zdroj, mail_id=mid):
                pridane += 1

        # ručná kontrola — dôvody sa spoja do jedného záznamu per mail
        dovody = []
        if not prijate and not neiste_ai:
            dovody.append("nič sa nedalo spárovať")
        if neiste_ai:
            dovody.append("AI neisto navrhuje: " + "; ".join(
                f"{it.id_hlasenia} → {it.status}" for it in neiste_ai))
        if len_heur:
            dovody.append("AI a heuristika sa nezhodli (našla len "
                          "heuristika): " + ", ".join(
                              it.id_hlasenia for it, _ in len_heur))
        if dovody:
            stavmod.pridaj_rucnu_kontrolu(
                data, servis_h, datum=datum, mail_id=mid,
                popis=f"mail z {datum} "
                      f"„{(meta.get('predmet') or '')[:60]}“: "
                      + "; ".join(dovody))

        stavove = (f"{pridane} poznámok" if prijate
                   else "RUČNÁ KONTROLA (0 spárovaných)")
        extra = f", {cudzie} neznámych čísel" if cudzie else ""
        extra += f", {len(neiste_ai)} AI-neistých" if neiste_ai else ""
        novy = "" if zaznamenane or pridane else " (už spracovaný)"
        print(f"  OK [{mid}] {nazov}: {stavove}{extra}{novy}")

    return data


# ---------- merge klasifikácií od agenta ----------

def _odpovedala_ina_vrstva(rec: dict, mail_id: str) -> bool:
    """True, ak k hláseniu z tohto mailu zapísala poznámku heuristika
    alebo AI vrstva (čokoľvek okrem zdroja 'agent')."""
    return any(p.get("mail_id") == mail_id
               and p["zdroj"] in ("telo", "priloha", "ai", "ai+heur")
               for p in rec["poznamky"])


def spracuj_klasifikacie(data: dict, mapa: dict) -> dict:
    """Zmerguje odpovede/<id>/klasifikacia.json (výstup agenta
    klasifikator-odpovedi, spúšťaný cez /klasifikuj-odpovede) do stavu:
    pokec ako poznámka (zdroj "agent", dedup mail_id+zdroj) a pole
    rec["klasifikacia"] (kategória/akcia/dobropis — kreslí ho excel
    a hárok Triáž; pri viacerých mailoch vyhráva najnovší).

    Validuje enumy — nevalidné položky sa zahadzujú a mail ide na ručnú
    kontrolu. Krížová kontrola: hlásenie, ktoré z tohto mailu nevidela
    heuristika ani AI vrstva, sa zapíše, ale mail sa flaguje (typicky
    verdikt z PDF/fotky, ktoré ostatné vrstvy nečítajú — info pre Evu,
    že to stojí na čítaní agenta)."""
    pocty: dict[str, int] = {}
    n_mailov = 0
    for mid, z in sorted(mapa.items(), key=lambda kv: (kv[1]["datum"], kv[0])):
        kl = z.get("klasifikacia")
        if not kl or z["servis_h"] is None or z["ndr"]:
            continue
        n_mailov += 1
        servis_h, datum = z["servis_h"], z["datum"]
        nazov = data["servisy"][servis_h]["nazov"]
        problemy, len_agent = [], []

        if kl.get("mail_typ") not in MAIL_TYPY:
            problemy.append(f"neznámy mail_typ {kl.get('mail_typ')!r}")

        for h in kl.get("hlasenia", []):
            hid = str(h.get("id_hlasenia", "")).strip()
            rec = data["hlasenia"].get(hid)
            kateg, akcia = h.get("kategoria"), h.get("akcia")
            conf = h.get("confidence")
            pokec = str(h.get("pokec") or "").strip()
            cislo = h.get("cislo_dobropisu") or None
            if (rec is None or not HLASENIE_RE.fullmatch(hid)
                    or kateg not in KATEGORIE
                    or akcia not in AKCIE_PORADIE
                    or conf not in CONFIDENCE_PORADIE or not pokec):
                problemy.append(
                    f"nevalidná položka {hid or '?'} "
                    f"({kateg}/{akcia}/{conf})")
                continue
            if kateg == "uzavrete-dobropisovane" and not cislo:
                # dobropis bez čísla nevie Eva zapísať — nech to pozrie
                akcia = "pozriet-rucne"
                if conf == "high":
                    conf = "medium"

            # prílohy hlásenia (balíček pre SAP) — cesty relatívne
            # k priečinku mailu, ukladajú sa s prefixom mail_id, lebo
            # hlásenie môže nazbierať prílohy z viacerých mailov
            prilohy_h = []
            for pr in (h.get("prilohy") or []):
                pr = str(pr).replace("\\", "/").strip().lstrip("/")
                if not pr:
                    continue
                if pr.startswith("_prepisy") or pr.endswith(".zip"):
                    continue  # prepisy a zipy do balíčka nepatria
                if not (z["dir"] / pr).exists() \
                        and (z["dir"] / "_rozbalene" / pr).exists():
                    # agenti občas vynechajú prefix _rozbalene/
                    pr = f"_rozbalene/{pr}"
                if (z["dir"] / pr).exists():
                    prilohy_h.append(f"{mid}/{pr}")
                else:
                    problemy.append(f"{hid}: priradená príloha "
                                    f"neexistuje: {pr}")

            stavmod.pridaj_odpoved(data, hid, text=pokec, confidence=conf,
                                   datum=datum, zdroj="agent", mail_id=mid)
            stara = rec.get("klasifikacia")
            zlucene = sorted(set(prilohy_h)
                             | set((stara or {}).get("prilohy", [])))
            if stara is None or stara.get("datum", "") <= datum:
                rec["klasifikacia"] = {
                    "kategoria": kateg, "akcia": akcia,
                    "cislo_dobropisu": cislo, "pokec": pokec,
                    "confidence": conf, "datum": datum, "mail_id": mid,
                    "prilohy": zlucene,
                }
            else:
                stara["prilohy"] = zlucene
            pocty[kateg] = pocty.get(kateg, 0) + 1
            if not _odpovedala_ina_vrstva(rec, mid):
                len_agent.append(hid)

        # opačný smer: iná vrstva z tohto mailu odpoveď našla,
        # ale agent hlásenie neklasifikoval
        kl_ids = {str(h.get("id_hlasenia", "")).strip()
                  for h in kl.get("hlasenia", [])}
        bez_klasifikacie = [
            hid for hid, rec in data["hlasenia"].items()
            if rec["servis_h"] == servis_h and hid not in kl_ids
            and _odpovedala_ina_vrstva(rec, mid)
        ]

        dovody = []
        if problemy:
            dovody.append("nevalidný výstup klasifikátora: "
                          + "; ".join(problemy))
        if len_agent:
            dovody.append("klasifikáciu má len agent (heuristika ani AI "
                          "to nevideli, typicky PDF/fotka): "
                          + ", ".join(len_agent))
        if bez_klasifikacie:
            dovody.append("bez klasifikácie, hoci odpoveď existuje: "
                          + ", ".join(bez_klasifikacie))
        if dovody:
            stavmod.pridaj_rucnu_kontrolu(
                data, servis_h, datum=datum, mail_id=mid,
                popis=f"klasifikátor, mail z {datum}: " + "; ".join(dovody))

        print(f"  KL [{mid}] {nazov}: {kl.get('mail_typ', '?')}, "
              f"{len(kl.get('hlasenia', []))} hlásení"
              + (f" ({len(problemy)} nevalidných)" if problemy else ""))

    if n_mailov:
        print(f"\nKlasifikácie: {n_mailov} mailov, hlásení podľa kategórií: "
              + (", ".join(f"{k}={n}" for k, n in sorted(pocty.items()))
                 or "žiadne"))
    else:
        print("\nŽiadne klasifikacia.json — klasifikáciu spúšťa "
              "/klasifikuj-odpovede v Claude Code.")
    return data


# ---------- balíčky príloh per reklamácia (prilohy/<hlasenie>/) ----------

def _syntetizuj_eml(z: dict) -> bytes:
    """Telo mailu ako .eml (len text + hlavičky, bez MIME príloh).

    Prílohy (xlsx, pdf…) sú v prilohy/<hid>/ aj ako samostatné súbory —
    vkladať ich aj do .eml by znamenalo každý súbor 2×."""
    from email.message import EmailMessage
    from email.utils import format_datetime
    from datetime import datetime

    meta = z["meta"]
    msg = EmailMessage()
    meno = meta.get("odosielatel_meno") or ""
    email_od = meta.get("odosielatel_email") or ""
    msg["From"] = (f"{meno} <{email_od}>" if email_od
                   else meta.get("odosielatel") or "neznámy")
    msg["Subject"] = meta.get("predmet") or "(bez predmetu)"
    if z["datum"]:
        try:
            msg["Date"] = format_datetime(
                datetime.fromisoformat(z["datum"]))
        except ValueError:
            pass
    msg.set_content(z["telo"] or "(bez tela)")
    return msg.as_bytes()


def generuj_prilohy(data: dict, mapa: dict) -> dict[str, str]:
    """Postaví prilohy/<hlasenie>/ nanovo (zmaže a pregeneruje — pohľad
    na cache + stav): kópia mailu ako .eml + súbory, ktoré hláseniu
    priradil klasifikátor. Eva priečinok len nahrá do SAP. Vráti mapu
    hlasenie -> relatívna cesta priečinka (excel na ňu linkuje)."""
    shutil.rmtree(PRILOHY_DIR, ignore_errors=True)
    dotknute = {hid: rec for hid, rec in data["hlasenia"].items()
                if rec.get("klasifikacia")}
    if not dotknute:
        return {}
    PRILOHY_DIR.mkdir()

    eml_cache: dict[str, bytes] = {}
    cesty: dict[str, str] = {}
    n_suborov = 0
    for hid, rec in sorted(dotknute.items()):
        kl = rec["klasifikacia"]
        maily = {kl.get("mail_id")}
        for pr in kl.get("prilohy", []):
            maily.add(pr.split("/", 1)[0])
        maily = {m for m in maily if m and m in mapa}
        if not maily:
            continue
        adresar = PRILOHY_DIR / hid
        adresar.mkdir()

        for m in sorted(maily):
            z = mapa[m]
            if m not in eml_cache:
                eml_cache[m] = _syntetizuj_eml(z)
            predmet = _bezpecny_nazov(
                z["meta"].get("predmet") or "bez predmetu", 60)
            ciel = _unikatna(
                adresar / f"{z['datum']} - mail - {predmet}.eml")
            ciel.write_bytes(eml_cache[m])
            n_suborov += 1

        for pr in kl.get("prilohy", []):
            m, _, rel = pr.partition("/")
            if m not in mapa or not rel:
                continue
            zdroj = mapa[m]["dir"] / rel
            if not zdroj.exists():
                continue
            sub = Path(rel)
            ciel = _unikatna(adresar / (
                f"{mapa[m]['datum']} - "
                f"{_bezpecny_nazov(sub.stem, 70)}{sub.suffix}"))
            shutil.copy2(zdroj, ciel)
            n_suborov += 1
        cesty[hid] = adresar.relative_to(ROOT).as_posix()

    print(f"Balíčky pre SAP: {PRILOHY_DIR.name}/ "
          f"({len(cesty)} reklamácií, {n_suborov} súborov)")
    return cesty


# ---------- strom priečinkov podľa servisov ----------

def _bezpecny_nazov(s: str, max_len: int = 70) -> str:
    s = re.sub(r'[\\/:*?"<>|\r\n\t]', "_", s)
    s = re.sub(r"\s+", " ", s).strip(" ._")
    return s[:max_len].rstrip(" ._") or "bez_nazvu"


def _unikatna(cesta: Path) -> Path:
    if not cesta.exists():
        return cesta
    n = 2
    while True:
        kand = cesta.with_name(f"{cesta.stem} ({n}){cesta.suffix}")
        if not kand.exists():
            return kand
        n += 1


def generuj_priecinky(data: dict, mapa: dict) -> dict[str, str]:
    """Postaví odpovede_podla_servisov/ nanovo (zmaže a pregeneruje — je to
    len pohľad na cache). Vráti mapu mail_id -> relatívna cesta k .txt
    (používa excel na odkazy pri ručnej kontrole)."""
    shutil.rmtree(STROM_DIR, ignore_errors=True)
    STROM_DIR.mkdir()
    cesty: dict[str, str] = {}

    for mid, z in sorted(mapa.items(), key=lambda kv: (kv[1]["datum"], kv[0])):
        meta, datum = z["meta"], z["datum"]
        if z["servis_h"] is None:
            adresar = STROM_DIR / "_nezaradene"
        else:
            adresar = STROM_DIR / _bezpecny_nazov(
                data["servisy"][z["servis_h"]]["nazov"])
        adresar.mkdir(exist_ok=True)

        meno, email = odosielatel_casti(meta)
        odkaz = owa_odkaz(meta)
        predmet = meta.get("predmet") or "(bez predmetu)"
        druh = "NEDORUCENE" if z["ndr"] else "mail"
        txt = _unikatna(adresar / f"{datum} - {druh} - "
                                  f"{_bezpecny_nazov(predmet, 60)}.txt")
        hlavicka = [
            f"Od:      {meno} <{email}>" if email else f"Od:      "
            f"{meta.get('odosielatel', '')}",
            f"Prijaté: {datum}",
            f"Predmet: {predmet}",
            ("V Outlooku: " + odkaz) if odkaz
            else f"V Outlooku: hľadaj podľa predmetu: {predmet}",
            "-" * 60,
            "",
        ]
        txt.write_text("\n".join(hlavicka) + z["telo"], encoding="utf-8")
        cesty[mid] = txt.relative_to(ROOT).as_posix()

        for nazov_prilohy in meta.get("prilohy", []):
            zdroj_subor = z["dir"] / nazov_prilohy
            if not zdroj_subor.exists():
                continue
            ciel = _unikatna(adresar / (
                f"{datum} - priloha - "
                f"{_bezpecny_nazov(zdroj_subor.stem, 70)}"
                f"{zdroj_subor.suffix}"))
            shutil.copy2(zdroj_subor, ciel)

    print(f"Strom priečinkov: {STROM_DIR.name}/ "
          f"({len(cesty)} mailov, {sum(1 for _ in STROM_DIR.iterdir())} "
          "priečinkov)")
    return cesty


# ---------- excel pohľad ----------

def _odpovedne_poznamky(poznamky) -> list:
    return [p for p in poznamky if p["zdroj"] in ODPOVED_ZDROJE]


def _text_odpovede(p: dict) -> str:
    text = p["text"]
    if p["zdroj"] == "ai":
        text = "[AI] " + text
    if p.get("confidence") == "medium":
        text += " ?"
    elif p.get("confidence") == "low":
        text += " ??"
    return text


def _historia(poznamky, mapa: dict) -> str:
    kusy = []
    for p in poznamky:
        if p["zdroj"] == "ziadost":
            continue  # úvodný záznam je všade rovnaký, v exceli len šumí
        zdroj = ZDROJ_POPIS.get(p["zdroj"], p["zdroj"])
        z = mapa.get(p.get("mail_id") or "")
        od = ""
        if z:
            _, email = odosielatel_casti(z["meta"])
            od = f" od {email or 'neznámy'}"
        kusy.append(f"{p['datum']} ({zdroj}{od}): {_text_odpovede(p)}")
    return " | ".join(kusy)


def _hyperlink(ws, row: int, col: int, text: str, ciel: str):
    # relatívne cesty musia byť platné URI (medzery, diakritika -> %xx),
    # inak Excel pri otvorení súbor „opraví" a všetky hyperlinky vyhodí
    if not ciel.startswith("http"):
        ciel = urllib.parse.quote(ciel, safe="/")
    c = ws.cell(row=row, column=col, value=text)
    c.hyperlink = ciel
    c.font = HYPERLINK_FONT


def generuj_excel(data: dict, mapa: dict, cesty: dict[str, str],
                  prilohy_cesty: dict[str, str]):
    dnes = date.today().isoformat()
    wb = openpyxl.load_workbook(SRC)
    ws = wb.active

    # Nájdi stĺpce Poznámka a Kategória; ak Kategória chýba, vlož ju na D=4
    _hlavicky = {(c.value or "").strip(): i for i, c in enumerate(ws[1], 1) if c.value}
    poznamka_col = _hlavicky.get("Poznámka", 3)
    if "Kategória" not in _hlavicky:
        ws.insert_cols(4)
        c_kat = ws.cell(1, 4, value="Kategória")
        c_kat.font = Font(bold=True)
        ws.column_dimensions["D"].width = 24
        kategoria_col = 4
    else:
        kategoria_col = _hlavicky["Kategória"]

    prvy = ws.max_column + 1
    for ofs, hlavicka in enumerate(NOVE_HLAVICKY):
        c = ws.cell(row=1, column=prvy + ofs, value=hlavicka)
        c.font = Font(bold=True)

    # servisy, ktoré majú aspoň jeden mail na ručnú kontrolu
    rucne_set = {h for h, srv in data["servisy"].items() if srv["rucna_kontrola"]}

    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        rec = data["hlasenia"].get(str(v).strip()) if v is not None else None
        if rec is None:
            continue  # NOSERVIS / servis bez emailu — žiadosť nešla
        hid = str(v).strip()
        if hid.isdigit():
            # zdroj má hlásenia ako text — číslo sa lepšie filtruje
            # a nesvieti zeleným rožkom
            ws.cell(row=row, column=1, value=int(hid))
        ws.cell(row=row, column=prvy + 0, value=rec["stav"])
        ws.cell(row=row, column=prvy + 1, value=rec["on_turn"])

        odpovede = _odpovedne_poznamky(rec["poznamky"])
        je_novy = False
        if odpovede:
            p = odpovede[-1]
            z = mapa.get(p.get("mail_id") or "")
            ws.cell(row=row, column=prvy + 2, value=_text_odpovede(p))
            ws.cell(row=row, column=prvy + 3, value=p["datum"])
            if z:
                meno, email = odosielatel_casti(z["meta"])
                ws.cell(row=row, column=prvy + 4,
                        value=f"{meno} ({email})" if meno else email)
                ws.cell(row=row, column=prvy + 5,
                        value=z["meta"].get("predmet") or "")
                odkaz = owa_odkaz(z["meta"])
                if odkaz:
                    _hyperlink(ws, row, prvy + 6, "otvoriť", odkaz)
                if z["meta"].get("stiahnute") == dnes:
                    je_novy = True
        if je_novy:
            for col in range(1, prvy + len(NOVE_HLAVICKY) + 1):
                ws.cell(row=row, column=col).fill = NOVA_FARBA
        ws.cell(row=row, column=prvy + 8,
                value=_historia(rec["poznamky"], mapa))
        if rec["servis_h"] in rucne_set:
            ws.cell(row=row, column=prvy + 9, value="X")

        kl = rec.get("klasifikacia")
        if kl:
            kateg = KATEGORIA_POPIS.get(kl["kategoria"], kl["kategoria"])
            if kl.get("confidence") == "medium":
                kateg += " ?"
            elif kl.get("confidence") == "low":
                kateg += " ??"
            ws.cell(row=row, column=poznamka_col, value=kl["pokec"])
            ws.cell(row=row, column=kategoria_col, value=kateg)
            ws.cell(row=row, column=prvy + 10, value=kateg)
            ws.cell(row=row, column=prvy + 11,
                    value=kl.get("cislo_dobropisu") or "")
            ws.cell(row=row, column=prvy + 12,
                    value=AKCIA_POPIS.get(kl["akcia"], kl["akcia"]))
            ws.cell(row=row, column=prvy + 13, value=kl["pokec"])
            if hid in prilohy_cesty:
                _hyperlink(ws, row, prvy + 14, "otvoriť",
                           prilohy_cesty[hid])

    sirky = [12, 8, 50, 14, 32, 46, 13, 10, 80, 16, 24, 16, 24, 70, 10]
    for ofs, w in enumerate(sirky):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(prvy + ofs)].width = w

    # ---- hárok Triáž (pracovný zoznam podľa akcií) ----
    if "Triáž" in wb.sheetnames:
        del wb["Triáž"]
    tz = wb.create_sheet("Triáž")
    tz.append(["Hlásenie", "Mail v Outlooku", "Prílohy", "Akcia",
               "Kategória", "Servis", "Číslo dobropisu", "Pokec pre SAP",
               "Dátum", "Priečinok servisu"])
    for c in tz[1]:
        c.font = Font(bold=True)

    triaz = sorted(
        ((hid, rec) for hid, rec in data["hlasenia"].items()
         if rec.get("klasifikacia")),
        key=lambda kv: (
            AKCIE_PORADIE.index(kv[1]["klasifikacia"]["akcia"]),
            kv[1]["klasifikacia"]["kategoria"],
            data["servisy"][kv[1]["servis_h"]]["nazov"].lower(),
            kv[0],
        ))
    riadok = 1
    for hid, rec in triaz:
        kl = rec["klasifikacia"]
        riadok += 1
        kateg = KATEGORIA_POPIS.get(kl["kategoria"], kl["kategoria"])
        if kl.get("confidence") == "medium":
            kateg += " ?"
        elif kl.get("confidence") == "low":
            kateg += " ??"
        tz.append([
            int(hid) if hid.isdigit() else hid, "", "",
            AKCIA_POPIS.get(kl["akcia"], kl["akcia"]), kateg,
            data["servisy"][rec["servis_h"]]["nazov"],
            kl.get("cislo_dobropisu") or "", kl["pokec"], kl["datum"], "",
        ])
        z = mapa.get(kl.get("mail_id") or "")
        if z:
            odkaz = owa_odkaz(z["meta"])
            if odkaz:
                _hyperlink(tz, riadok, 2, "otvoriť", odkaz)
        if hid in prilohy_cesty:
            _hyperlink(tz, riadok, 3, "otvoriť", prilohy_cesty[hid])
    for col, w in zip("ABCDEFGHIJ",
                      [12, 14, 10, 24, 24, 36, 16, 80, 11, 14]):
        tz.column_dimensions[col].width = w

    # ---- hárok Prehľad servisov ----
    if "Prehľad servisov" in wb.sheetnames:
        del wb["Prehľad servisov"]
    ps = wb.create_sheet("Prehľad servisov")
    ps.append(["Servis", "Emaily", "Odpovedal", "Posledná odpoveď",
               "Posledný mail (predmet)", "Priečinok", "Hlásení",
               "Zodpovedaných", "Ručná kontrola", "Nedoručené"])
    for c in ps[1]:
        c.font = Font(bold=True)

    pocty = {}
    for rec in data["hlasenia"].values():
        c = pocty.setdefault(rec["servis_h"], [0, 0])
        c[0] += 1
        if rec["stav"] != "odoslane":
            c[1] += 1

    # servis -> posledný mail (datum, mail_id)
    posledny_mail = {}
    for mid, z in mapa.items():
        h = z["servis_h"]
        if h and not z["ndr"]:
            if h not in posledny_mail or z["datum"] > posledny_mail[h][0]:
                posledny_mail[h] = (z["datum"], mid)

    odpovedalo = 0
    riadok = 1
    for h, srv in sorted(data["servisy"].items(),
                         key=lambda kv: kv[1]["nazov"].lower()):
        riadok += 1
        celkom, zodp = pocty.get(h, [0, 0])
        datumy = [x["datum"] for x in srv["odpovedal"]]
        if datumy:
            odpovedalo += 1
        emaily = srv["email1"] + (f"; {srv['email2']}" if srv["email2"] else "")
        predmet = ""
        if h in posledny_mail:
            predmet = (mapa[posledny_mail[h][1]]["meta"].get("predmet")
                       or "")[:80]
        ps.append([
            srv["nazov"], emaily,
            "áno" if datumy else "nie",
            max(datumy) if datumy else "",
            predmet, "",
            celkom, zodp,
            len(srv["rucna_kontrola"]) or "",
            len(srv.get("nedorucene", [])) or "",
        ])
    for col, w in zip("ABCDEFGHIJ", [40, 45, 10, 16, 50, 10, 9, 14, 14, 12]):
        ps.column_dimensions[col].width = w

    try:
        wb.save(VYSTUP_XLSX)
    except PermissionError:
        print(f"\nPOZOR: {VYSTUP_XLSX.name} je otvorený v Exceli — "
              "zatvor ho a spusti znova (stačí --len-excel).")
        return

    mlcia = len(data["servisy"]) - odpovedalo
    ndr_srv = [s["nazov"] for s in data["servisy"].values()
               if s.get("nedorucene")]
    print(f"\nExcel: {VYSTUP_XLSX.name}")
    print(f"Servisov odpovedalo {odpovedalo}/{len(data['servisy'])}, "
          f"mlčí {mlcia}.")
    if ndr_srv:
        print(f"Nedoručiteľné adresy ({len(ndr_srv)}): "
              + "; ".join(ndr_srv))
    nez = data.get("nezaradene", [])
    if nez:
        print(f"Nezaradené maily (nespárované so servisom): {len(nez)}")


def main():
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    parser.add_argument("--len-excel", action="store_true",
                        help="Bez spracovania mailov (bez AI) — len strom "
                             "priečinkov + excel zo stavového súboru")
    parser.add_argument("--len-priecinky", action="store_true",
                        help="Iba pregeneruje strom priečinkov")
    parser.add_argument("--ai-znova", action="store_true",
                        help="Ignoruje AI cache a číta maily znova")
    parser.add_argument("--len-klasifikacie", action="store_true",
                        help="Len merge klasifikacia.json do stavu "
                             "(bez AI vrstvy) + strom a excel")
    args = parser.parse_args()

    data = stavmod.load_stav()
    mapa = zostav_mapu_mailov(data)
    if args.len_klasifikacie:
        data = spracuj_klasifikacie(data, mapa)
        stavmod.save_stav(data)
    elif not (args.len_excel or args.len_priecinky):
        data = spracuj_maily(data, mapa, ai_znova=args.ai_znova)
        data = spracuj_klasifikacie(data, mapa)
        stavmod.save_stav(data)
    shutil.rmtree(STROM_DIR, ignore_errors=True)
    cesty = {}
    prilohy_cesty = generuj_prilohy(data, mapa)
    if not args.len_priecinky:
        generuj_excel(data, mapa, cesty, prilohy_cesty)


if __name__ == "__main__":
    main()
