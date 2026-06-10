"""
Rozdelí zdrojový excel `dpb servis 180-360.xlsx` na samostatné súbory podľa servisu
(stĺpec H zdroja). Každý výstupný súbor obsahuje len riadky pre daný servis a má
stĺpce: Hlášení, Krát.text materiálu, Název značky, Servis, Datum hlášení,
Poznámka servis (posledný stĺpec je prázdny — vyplní ho servis).

Spustenie testu (vytvorí len vybrané servisy):
    python rozdel_servisy.py --test

Spustenie celého rozdelenia:
    python rozdel_servisy.py
"""
import sys
import re
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parent
SRC = ROOT / "dpb servis 180-360.xlsx"
OUT_DIR = ROOT / "servis_subory"

# Indexy stĺpcov v zdroji (0-based): A=0, E=4, F=5, H=7, L=11
SRC_COLS = [0, 4, 5, 7, 11]
HEADERS = ["Hlášení", "Krát.text materiálu", "Název značky", "Servis",
           "Datum hlášení", "Poznámka servis"]

INVALID_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|]')


def sanitize_filename(name: str) -> str:
    name = INVALID_FILENAME_CHARS.sub("_", name)
    # ponechávame trailing bodku (napr. "s.r.o."), strip len medzery
    return name.strip()


def extract_service_name(h_value: str) -> str:
    """Z '3000008563 - Candy Hoover ČR s.r.o.' vráti 'Candy Hoover ČR s.r.o.'."""
    if " - " in h_value:
        return h_value.split(" - ", 1)[1].strip()
    return h_value.strip()


def extract_service_code(h_value: str) -> str:
    if " - " in h_value:
        return h_value.split(" - ", 1)[0].strip()
    return ""


def load_groups():
    """Načíta zdroj a vráti dict: h_value -> list[row_values_for_output]."""
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.active
    groups = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        h = row[7]
        if not h:
            continue
        if str(h).startswith("NOSERVIS"):
            continue
        out_row = [row[i] for i in SRC_COLS] + [None]  # posledný stĺpec prázdny
        # Hlášení (stĺpec A) skonvertuj na int – v zdroji je ako text
        if out_row[0] is not None:
            try:
                out_row[0] = int(out_row[0])
            except (ValueError, TypeError):
                pass  # ak by sa nedalo, nechaj pôvodné
        groups[h].append(out_row)
    return groups


def build_filename_map(groups):
    """Vráti dict: h_value -> filename (bez prípony). Pri kolíziách názvu
    pridá kód do zátvorky, ostatné názvy nechá čisté."""
    by_name = defaultdict(list)
    for h in groups:
        by_name[extract_service_name(h)].append(h)

    out = {}
    for name, h_list in by_name.items():
        safe_name = sanitize_filename(name)
        if len(h_list) == 1:
            out[h_list[0]] = safe_name
        else:
            for h in h_list:
                code = extract_service_code(h)
                out[h] = f"{safe_name} ({code})"
    return out


def write_service_file(filepath: Path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADERS)
    # hlavička – jemne tučná, aby sa odlíšila (vzor mal bez bold, ale dáta sú
    # ľahšie čitateľné s tučnou hlavičkou; ak nechceš, riadok nižšie zmaž)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(row)

    # formát dátumu pre stĺpec E
    for row_cells in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row_cells:
            if cell.value is not None:
                cell.number_format = "d.m.yyyy"

    # rozumné šírky stĺpcov
    widths = {"A": 14, "B": 45, "C": 18, "D": 38, "E": 14, "F": 40}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    wb.save(filepath)


def main():
    test_mode = "--test" in sys.argv
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    groups = load_groups()
    name_map = build_filename_map(groups)

    # v test móde len 2 servisy: Candy Hoover ČR s.r.o. a Mc TREE a.s.
    if test_mode:
        wanted = {"3000008563 - Candy Hoover ČR s.r.o.",
                  "3000003226 - Mc TREE a.s."}
        targets = {h: rows for h, rows in groups.items() if h in wanted}
    else:
        targets = groups

    print(f"Spracovavam {len(targets)} servisov "
          f"(test_mode={test_mode}, celkom v zdroji={len(groups)})")
    failed = []
    for h, rows in targets.items():
        fname = name_map[h] + ".xlsx"
        path = OUT_DIR / fname
        try:
            write_service_file(path, rows)
            print(f"  OK {fname}  ({len(rows)} riadkov)")
        except PermissionError:
            failed.append(fname)
            print(f"  ZAMKNUTY (otvoreny v Exceli?): {fname}")

    print(f"\nHotovo. Vystup v: {OUT_DIR}")
    if failed:
        print(f"\nPOZOR – nepodarilo sa zapisat {len(failed)} suborov "
              f"(zatvor ich v Exceli a spusti znova):")
        for f in failed:
            print(f"  - {f}")


if __name__ == "__main__":
    main()
