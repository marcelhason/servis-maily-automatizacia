"""
Parsovanie odpovedí od servisov — vymeniteľné rozhranie.

Verejné rozhranie (na ktoré sa viaže spracuj_odpovede.py):

    parse_reply(text, zname_hlasenia)      -> list[ParsedItem]   # telo mailu
    parse_xlsx_reply(path, zname_hlasenia) -> list[ParsedItem]   # vrátená príloha
    parse_reply_ai(telo, zname_hlasenia, prilohy_text)           # Claude API
    zluc_ai_heur(ai, heur, min_confidence) -> ZlucenyVysledok

Dve vrstvy:
  HEURISTIKA (parse_reply) — deterministická, zadarmo: v texte hľadá čísla
  Hlášení (10-miestne, prefix 880) a priraďuje k nim text riadku. Nič neháda.
  AI (parse_reply_ai) — Claude číta CELÝ mail (vrátane citovanej tabuľky,
  kam servisy často vpisujú odpovede) + texty xlsx príloh. Vyžaduje
  ANTHROPIC_API_KEY v config.py; bez neho ai_dostupne() vráti False.

Zlúčenie (zluc_ai_heur): zhoda AI+heuristiky -> confidence high; len-AI
sa prijme len od prahu AI_MIN_CONFIDENCE (config.py), pod ním ide návrh
na ručnú kontrolu; len-heuristika sa prijme, ale mail sa tiež flaguje
(AI to mala vidieť) — žiadne tiché hádanie.

POZOR na citácie (heuristika): odpoveď typicky cituje náš pôvodný mail,
ktorý obsahuje tabuľku so VŠETKÝMI hláseniami daného servisu. Bez orezania
citácie by heuristika „spárovala" úplne všetko, preto parse_reply najprv
orezáva text (orez_citaciu). AI naopak dostáva text NEorezaný.

Samotest:  python parsers.py
"""
import json
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

try:
    import config as _cfg
except ImportError:  # bez lokálneho configu bežia vzorové hodnoty
    import config_example as _cfg

CONFIDENCE_PORADIE = ("low", "medium", "high")

# Čísla Hlášení v zdroji: 10-miestne, začínajú 880 (napr. 8803000002)
HLASENIE_RE = re.compile(r"\b880\d{7}\b")

# Pôvodné hlavičky príloh (rozdel_servisy.HEADERS) — bunky v týchto stĺpcoch
# NIE sú odpoveď servisu. Všetko ostatné (Poznámka servis aj prípadné nové
# stĺpce, ktoré si servis pridal) sa berie ako odpoveď.
_ZNAME_HLAVICKY = ("hlaseni", "krat.text materialu", "nazev znacky",
                   "servis", "datum hlaseni")

# Oddeľovače, ktoré po odstránení čísla hlásenia z riadku nemajú zostať
# trčať na krajoch statusu.
_OKRAJE = " \t-–—:•|,;.…"


@dataclass
class ParsedItem:
    id_hlasenia: str
    status: str
    confidence: str  # enum CONFIDENCE_PORADIE; heuristika dáva high (číslo
    #                  patrí servisu) / low (cudzie), AI aj medium (odvodené)


def confidence_aspon(c: str, prah: str) -> bool:
    """True, ak je confidence c aspoň na úrovni prahu (low<medium<high)."""
    return CONFIDENCE_PORADIE.index(c) >= CONFIDENCE_PORADIE.index(prah)


def _norm(s: str) -> str:
    """Lowercase bez diakritiky — na porovnávanie hlavičiek a markerov."""
    nfd = unicodedata.normalize("NFD", s)
    return "".join(c for c in nfd if not unicodedata.combining(c)).lower()


# ---------- orezanie citácie ----------

# Riadok, ktorým začína citovaný pôvodný mail (porovnáva sa _norm verzia).
_CITACIA_RIADOK = re.compile(
    r"^\s*(>|_{6,}|(od|from)\s*:\s|-{2,}\s*original message)")
_CITACIA_FRAZY = (
    "ziadost o informacie k dorieseniu reklamacii",  # predmet/nadpis citácie
    "obraciam sa na vas s prosbou",                  # intro nášho mailu
)


def orez_citaciu(text: str) -> str:
    """Vráti text po prvý riadok, ktorý vyzerá ako začiatok citácie
    pôvodného mailu (hlavička Od:/From:, oddeľovač, náš predmet/intro)."""
    lines = text.splitlines()
    for i, line in enumerate(lines):
        n = _norm(line)
        if _CITACIA_RIADOK.match(n):
            return "\n".join(lines[:i])
        if any(f in n for f in _CITACIA_FRAZY):
            return "\n".join(lines[:i])
    return text


# ---------- heuristický parser tela mailu ----------

def parse_reply_heuristic(text: str, zname_hlasenia: set[str]
                          ) -> list[ParsedItem]:
    """Nájde čísla hlásení a priradí k nim text riadku, na ktorom stoja.
    Ak je číslo na riadku samo, vezme najbližší nasledujúci neprázdny riadok
    bez čísla. Viac výskytov toho istého čísla sa zlúči do jedného statusu."""
    text = orez_citaciu(text)
    lines = text.splitlines()
    statusy: dict[str, list[str]] = {}

    for i, line in enumerate(lines):
        cisla = HLASENIE_RE.findall(line)
        if not cisla:
            continue
        status = re.sub(r"\s{2,}", " ",
                        HLASENIE_RE.sub("", line)).strip(_OKRAJE)
        if not status:
            for nasl in lines[i + 1:]:
                nasl = nasl.strip()
                if not nasl:
                    continue
                if HLASENIE_RE.search(nasl):
                    break  # ďalšie číslo — tento riadok patrí jemu
                status = nasl.strip(_OKRAJE)
                break
        if not status:
            status = "(odpoveď bez textu pri čísle — pozri celý mail)"
        for c in cisla:
            if status not in statusy.setdefault(c, []):
                statusy[c].append(status)

    return [
        ParsedItem(
            id_hlasenia=c,
            status="; ".join(texty),
            confidence="high" if c in zname_hlasenia else "low",
        )
        for c, texty in statusy.items()
    ]


# Aktívna implementácia — sem sa neskôr dá dosadiť AI variant.
parse_reply = parse_reply_heuristic


# ---------- parser vrátenej xlsx prílohy ----------

def _fmt_bunka(v) -> str:
    import datetime as dt
    if isinstance(v, (dt.datetime, dt.date)):
        return f"{v.day}.{v.month}.{v.year}"
    return str(v).strip()


def parse_xlsx_reply(path: Path, zname_hlasenia: set[str]
                     ) -> list[ParsedItem]:
    """Prečíta servisom vrátenú prílohu. Stĺpec Hlášení nájde podľa hlavičky
    (fallback: stĺpec, ktorého hodnoty sedia na 880-vzor). Ako odpoveď berie
    všetky neprázdne bunky v stĺpcoch MIMO pôvodných dátových hlavičiek —
    teda Poznámka servis aj prípadné stĺpce, ktoré si servis pridal."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    hlavicky = [(_norm(str(c)).strip() if c is not None else "")
                for c in rows[0]]

    idx_hlasenie = None
    for i, h in enumerate(hlavicky):
        if "hlasen" in h:
            idx_hlasenie = i
            break
    if idx_hlasenie is None:
        for i in range(len(rows[0])):
            if any(r[i] is not None and HLASENIE_RE.fullmatch(str(r[i]).strip())
                   for r in rows[1:]):
                idx_hlasenie = i
                break
    if idx_hlasenie is None:
        return []

    if any(h for h in hlavicky):
        # presná zhoda — substring by vylúčil aj "poznamka servis" (obsahuje
        # "servis"), a to je práve stĺpec s odpoveďou
        idx_odpovede = [i for i, h in enumerate(hlavicky)
                        if i != idx_hlasenie and h not in _ZNAME_HLAVICKY]
    else:
        idx_odpovede = list(range(5, len(rows[0])))

    out = []
    for r in rows[1:]:
        if idx_hlasenie >= len(r) or r[idx_hlasenie] is None:
            continue
        m = HLASENIE_RE.search(str(r[idx_hlasenie]).strip())
        if not m:
            continue
        cislo = m.group(0)
        kusy = [_fmt_bunka(r[i]) for i in idx_odpovede
                if i < len(r) and r[i] is not None
                and str(r[i]).strip() != ""]
        if not kusy:
            continue
        out.append(ParsedItem(
            id_hlasenia=cislo,
            status=" | ".join(kusy),
            confidence="high" if cislo in zname_hlasenia else "low",
        ))
    return out


# ---------- AI parser (Claude API) ----------

_AI_SYSTEM = """\
Čítaš odpoveď servisnej firmy na našu žiadosť o stav reklamácií spotrebičov.
Každá reklamácia má 10-miestne číslo hlásenia začínajúce 880. Telo mailu môže
obsahovať citáciu našej pôvodnej žiadosti s tabuľkou hlásení — servisy často
vpisujú odpovede priamo do nej. Dostaneš aj textový prepis xlsx príloh
a zoznam hlásení, ktoré tomuto servisu patria.

Vráť POLE JSON objektov {"id": "...", "status": "...", "confidence": "..."}:
- id: číslo hlásenia, ku ktorému sa servis vecne vyjadril,
- status: stručný stav v jazyku mailu (napr. "vybavené dobropisom č. 123",
  "diel objednaný, oprava do 20.6.") — výhradne fakty z mailu, nedomýšľaj,
- confidence: "high" = explicitná odpoveď pri konkrétnom čísle,
  "medium" = odvodené (kolektívna odpoveď rozpísaná na hlásenia),
  "low" = neisté alebo nejednoznačné.

Pravidlá:
- Kolektívnu odpoveď platnú pre všetky reklamácie ("všetky sú vybavené")
  rozpíš na všetky známe hlásenia servisu s confidence "medium".
- Riadok citovanej tabuľky, do ktorého servis NIČ nedopísal, nie je odpoveď.
- Automatické potvrdenky ticketov, "preveríme a ozveme sa", nedoručenky
  a maily bez vecnej informácie => prázdne pole [].
Vráť IBA JSON pole, žiadny iný text."""


def ai_dostupne() -> bool:
    return bool(getattr(_cfg, "ANTHROPIC_API_KEY", ""))


def xlsx_na_text(path: Path, max_riadkov: int = 200) -> str:
    """Prepíše hárok xlsx prílohy na text (riadky oddelené \\n, bunky ' | ')
    pre AI prompt."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    riadky = []
    for r in ws.iter_rows(values_only=True):
        if all(v is None for v in r):
            continue
        riadky.append(" | ".join(_fmt_bunka(v) if v is not None else ""
                                 for v in r).rstrip(" |"))
        if len(riadky) >= max_riadkov:
            riadky.append(f"... (skrátené na {max_riadkov} riadkov)")
            break
    return "\n".join(riadky)


def eml_na_text(path: Path) -> str:
    """Prepíše .eml prílohu (preposlaný mail) na text pre AI prompt."""
    import email
    from email import policy

    msg = email.message_from_bytes(path.read_bytes(), policy=policy.default)
    casti = [f"Od: {msg.get('From', '')}",
             f"Predmet: {msg.get('Subject', '')}", ""]
    try:
        telo = msg.get_body(preferencelist=("plain", "html"))
        if telo is not None:
            obsah = telo.get_content()
            if telo.get_content_type() == "text/html":
                obsah = re.sub(r"<[^>]+>", " ", obsah)
            casti.append(re.sub(r"\n{3,}", "\n\n", obsah).strip())
    except Exception:
        casti.append("(telo sa nepodarilo dekódovať)")
    return "\n".join(casti)


def parse_reply_ai(telo: str, zname_hlasenia: set[str],
                   prilohy_text: str = "") -> list[ParsedItem]:
    """Prečíta mail Claudom. Telo sa posiela NEorezané (aj s citáciou —
    odpovede bývajú vpísané v citovanej tabuľke). Vyhadzuje výnimky API
    nahor — caller rozhoduje, čo s neprečítaným mailom."""
    # antivírus/proxy podpisuje TLS vlastným certom, ktorý nie je v certifi
    # -> dôveruj certifikátom z Windows úložiska
    import truststore
    truststore.inject_into_ssl()
    import anthropic

    casti = [
        "Hlásenia patriace tomuto servisu:\n"
        + ", ".join(sorted(zname_hlasenia)),
        "TELO MAILU:\n" + telo.strip(),
    ]
    if prilohy_text.strip():
        casti.append("PREPIS PRÍLOH:\n" + prilohy_text.strip())

    client = anthropic.Anthropic(api_key=_cfg.ANTHROPIC_API_KEY)
    resp = client.messages.create(
        model=getattr(_cfg, "AI_MODEL", "claude-sonnet-4-6"),
        max_tokens=8000,
        system=_AI_SYSTEM,
        messages=[{"role": "user", "content": "\n\n".join(casti)}],
    )
    surove = "".join(b.text for b in resp.content if b.type == "text")

    # model má vrátiť len JSON pole, ale keby okolo pridal text/```,
    # vezmi obsah od prvej [ po poslednú ]
    zac, kon = surove.find("["), surove.rfind("]")
    if zac < 0 or kon <= zac:
        raise ValueError(f"AI nevrátila JSON pole: {surove[:200]!r}")
    polozky = json.loads(surove[zac:kon + 1])

    out, videne = [], set()
    for p in polozky:
        if not isinstance(p, dict):
            continue
        pid = str(p.get("id", "")).strip()
        status = str(p.get("status", "")).strip()
        conf = str(p.get("confidence", "")).strip().lower()
        if (not HLASENIE_RE.fullmatch(pid) or not status
                or conf not in CONFIDENCE_PORADIE or pid in videne):
            continue  # nevalidná položka sa zahadzuje, nehádame
        videne.add(pid)
        out.append(ParsedItem(id_hlasenia=pid, status=status, confidence=conf))
    return out


# ---------- zlúčenie AI × heuristika ----------

@dataclass
class ZlucenyVysledok:
    # poznámky na zapísanie do stavu: [(ParsedItem, zdroj)]
    prijate: list = field(default_factory=list)
    # len-AI návrhy POD prahom — nezapisujú sa, idú do ručnej kontroly
    neiste_ai: list = field(default_factory=list)
    # heuristika našla, AI nie (nezhoda) — zapísané sú, ale mail sa flaguje
    len_heur: list = field(default_factory=list)


def zluc_ai_heur(ai_items: list[ParsedItem],
                 heur_polozky: list[tuple[ParsedItem, str]],
                 min_confidence: str | None = None) -> ZlucenyVysledok:
    """Pravidlá (per číslo hlásenia):
    - AI aj heuristika -> jedna poznámka, status z AI, confidence high,
      zdroj "ai+heur",
    - len AI s confidence >= prahu -> zdroj "ai",
    - len AI pod prahom -> neiste_ai (caller flaguje ručnú kontrolu),
    - len heuristika -> poznámka s pôvodným zdrojom (telo/priloha)
      + len_heur (caller flaguje ručnú kontrolu — AI to mala vidieť)."""
    if min_confidence is None:
        min_confidence = getattr(_cfg, "AI_MIN_CONFIDENCE", "medium")
    v = ZlucenyVysledok()
    heur_ids = {it.id_hlasenia for it, _ in heur_polozky}
    ai_ids = {it.id_hlasenia for it in ai_items}

    for it in ai_items:
        if it.id_hlasenia in heur_ids:
            v.prijate.append((ParsedItem(it.id_hlasenia, it.status, "high"),
                              "ai+heur"))
        elif confidence_aspon(it.confidence, min_confidence):
            v.prijate.append((it, "ai"))
        else:
            v.neiste_ai.append(it)

    for it, zdroj in heur_polozky:
        if it.id_hlasenia not in ai_ids:
            v.prijate.append((it, zdroj))
            v.len_heur.append((it, zdroj))
    return v


# ---------- samotest ----------

def _selftest():
    zname = {"8803000002", "8803000001"}

    text = (
        "Dobrý deň,\n"
        "posielame stav reklamácií:\n"
        "8803000002 - vybavené dobropisom č. 2026/0042\n"
        "8803000001\n"
        "  oprava dokončená, odoslané 5.6.2026 na centrálny sklad\n"
        "8809999999: toto číslo nie je naše\n"
        "\n"
        "S pozdravom\n"
        "Od: Lucia XY\n"
        "Predmet: Žiadosť o informácie k doriešeniu reklamácií – ...\n"
        "8803000002 citovaný riadok z tabuľky, NESMIE sa parsovať druhýkrát\n"
    )
    vysledky = {p.id_hlasenia: p for p in parse_reply(text, zname)}

    assert set(vysledky) == {"8803000002", "8803000001", "8809999999"}, vysledky
    assert vysledky["8803000002"].status == "vybavené dobropisom č. 2026/0042"
    assert vysledky["8803000002"].confidence == "high"
    assert vysledky["8803000001"].status.startswith("oprava dokončená")
    assert vysledky["8809999999"].confidence == "low"

    # citácia: text za "Od:" sa nesmie dostať do statusov
    assert "citovaný" not in vysledky["8803000002"].status

    # orezanie podľa intro frázy (bez diakritiky v teste schválne)
    t2 = "8803000002 OK\nObraciam sa na Vas s prosbou...\n8803000001 X\n"
    v2 = parse_reply(t2, zname)
    assert len(v2) == 1 and v2[0].id_hlasenia == "8803000002", v2

    # mail bez čísel -> prázdny list (ručnú kontrolu rieši volajúci)
    assert parse_reply("Všetky vaše reklamácie sú vybavené.", zname) == []

    # ---- zlúčenie AI × heuristika ----
    assert confidence_aspon("medium", "medium")
    assert confidence_aspon("high", "medium")
    assert not confidence_aspon("low", "medium")

    ai = [ParsedItem("8803000002", "vybavené dobropisom", "high"),   # zhoda
          ParsedItem("8803000001", "všetky vybavené", "medium"),     # len AI, nad prahom
          ParsedItem("8801111111", "asi vybavené?", "low")]          # len AI, pod prahom
    heur = [(ParsedItem("8803000002", "vybavené dobropisom č. 42", "high"),
             "priloha"),
            (ParsedItem("8802222222", "oprava hotová", "high"), "telo")]  # len heur
    z = zluc_ai_heur(ai, heur, min_confidence="medium")

    podla_id = {it.id_hlasenia: (it, zdroj) for it, zdroj in z.prijate}
    assert podla_id["8803000002"][1] == "ai+heur"
    assert podla_id["8803000002"][0].confidence == "high"
    assert podla_id["8803000001"][1] == "ai"
    assert podla_id["8802222222"][1] == "telo"
    assert [it.id_hlasenia for it in z.neiste_ai] == ["8801111111"]
    assert [it.id_hlasenia for it, _ in z.len_heur] == ["8802222222"]
    assert "8801111111" not in podla_id  # pod prahom sa nezapisuje

    print("Samotest parse_reply + zluc_ai_heur: OK")


if __name__ == "__main__":
    _selftest()
